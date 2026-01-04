
import pandas as pd
import sys
from openpyxl.utils import get_column_letter

# Set encoding for output
sys.stdout.reconfigure(encoding='utf-8')

excel_path = 'd:/Workspace/Manage_TO/TO_Table.xlsx'
csv_path = 'd:/Workspace/Manage_TO/TO_list.csv'
sheet_name = '2-본부'

try:
    # 1. Read Excel Headers to get Order
    print("Reading Excel Headers...")
    df_ex = pd.read_excel(excel_path, sheet_name=sheet_name, header=None)
    
    # Range 1: Q(16) to CB(79). Row 13 (Index 12).
    # Range 2: CC(80) to CI(86). Row 10 (Index 9).
    
    rank_order = []
    
    # Q is Index 16. CB is 79.
    for col_idx in range(16, 80): # 16 to 79 inclusive
        rank_name = str(df_ex.iloc[12, col_idx]).replace('\n', ' ').strip()
        col_letter = get_column_letter(col_idx + 1)
        rank_order.append({'col': col_letter, 'rank': rank_name})
        
    # CC is Index 80. CI is 86.
    for col_idx in range(80, 87):
        rank_name = str(df_ex.iloc[9, col_idx]).replace('\n', ' ').strip()
        col_letter = get_column_letter(col_idx + 1)
        rank_order.append({'col': col_letter, 'rank': rank_name})
        
    # 2. Read CSV to get Counts
    print("Reading CSV Data...")
    df_csv = pd.read_csv(csv_path, encoding='utf-8-sig')
    general_df = df_csv[df_csv['공무원_종류'] == '일반직']
    
    # Count by Rank
    rank_counts = general_df['직급'].value_counts().to_dict()
    
    # 3. Combine and Print
    print("\n\n--- General Service Ranks (Ordered by Column) ---")
    print(f"{'Col':<5} | {'Count':<6} | {'Rank Name'}")
    print("-" * 60)
    
    total_check = 0
    for item in rank_order:
        r_name = item['rank']
        col = item['col']
        cnt = rank_counts.get(r_name, 0)
        
        # Only print if relevant? Or print all zero counts too?
        # User asked for order, implying "all columns". 
        # But if count is 0, maybe less relevant?
        # Let's print if count > 0 OR if it's the requested range.
        if cnt >= 0:
             print(f"{col:<5} | {cnt:<6} | {r_name}")
             total_check += cnt
             
    print("-" * 60)
    print(f"Total Count from Table: {total_check}")
    
except Exception as e:
    print(f"Error: {e}")
